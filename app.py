from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
import os
import mysql.connector
from mysql.connector import Error as MySQLError
from openpyxl import Workbook
from datetime import datetime
import io

app = Flask(__name__)

# Configure MySQL
app.config['MYSQL_HOST'] = 'database-1.cxuy6moa43xw.us-east-1.rds.amazonaws.com'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = 'Daniel#MYSQL'
app.config['MYSQL_DB'] = 'software_contest'

# Email Configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'peccccsoftwarecontest@gmail.com'  # Replace with your email
app.config['MAIL_PASSWORD'] = 'idrx oanp rccl sbyh'     # Replace with your app password
app.config['MAIL_DEFAULT_SENDER'] = 'peccccsoftwarecontest@gmail.com'  # Replace with your email

# Secret key for session and flash messages
app.secret_key = 'your_secret_key_here'  # Replace with a secure secret key

# Initialize Flask-Mail extension
mail = Mail(app)

# Function to get MySQL connection
def get_db_connection():
    return mysql.connector.connect(
        host=app.config['MYSQL_HOST'],
        user=app.config['MYSQL_USER'],
        password=app.config['MYSQL_PASSWORD'],
        database=app.config['MYSQL_DB']
    )


@app.route('/')
def index():
    return render_template('rules.html')

@app.route('/registration')
def registration():
    return render_template('index.html')

@app.route('/success')
def success():
    # Get the success message from session
    success_message = session.pop('success_message', None)
    if not success_message:
        return redirect(url_for('registration'))
    return render_template('success.html', message=success_message)

@app.route('/register', methods=['POST'])
def register():
    if request.method == 'POST':
        # Get form data
        name = request.form['name']
        roll_number = request.form['rollNumber'].upper()  # Convert to uppercase
        register_number = request.form['registerNumber']
        email = request.form['email']
        department = request.form['department']

        try:
            # Get database connection
            conn = get_db_connection()
            cur = conn.cursor(dictionary=True)

            # Check for existing entries
            check_query = """
                SELECT * FROM Students 
                WHERE rollNumber = %s 
                OR registerNumber = %s 
                OR email = %s
            """
            cur.execute(check_query, (roll_number, register_number, email))
            existing_entry = cur.fetchone()
            
            if existing_entry:
                cur.close()
                conn.close()
                if existing_entry['rollNumber'] == roll_number:
                    flash('Roll number already registered. Please use a different roll number.', 'error')
                elif existing_entry['registerNumber'] == register_number:
                    flash('Register number already registered. Please use a different register number.', 'error')
                elif existing_entry['email'] == email:
                    flash('Email address already registered. Please use a different email address.', 'error')
                return redirect(url_for('registration'))

            # If no duplicates found, proceed with insertion
            insert_query = """
                INSERT INTO Students (name, rollNumber, registerNumber, email, department)
                VALUES (%s, %s, %s, %s, %s)
            """
            cur.execute(insert_query, (name, roll_number, register_number, email, department))
            conn.commit()

            # Send confirmation email
            try:
                msg = Message(
                    'Registration Successful - Panimalar Engineering College',
                    recipients=[email]
                )
                msg.body = f"""
Dear {name},

Thank you for registering for the Software Contest at Panimalar Engineering College.

Your registration details:
- Name: {name}
- Roll Number: {roll_number}
- Register Number: {register_number}
- Department: {department}

We look forward to seeing you at the contest!

Best regards,
Panimalar Engineering College
                """
                mail.send(msg)
                success_message = f"Registration successful! A confirmation email has been sent to {email}. If you don't see the message in your inbox, please check your spam folder."
            except Exception as e:
                print(f"Failed to send email: {str(e)}")
                success_message = 'Registration successful! However, we could not send the confirmation email. Please save your registration details.'

            # Close connection
            cur.close()
            conn.close()
            
            # Store success message in session and redirect to success page
            session['success_message'] = success_message
            return redirect(url_for('success'))

        except MySQLError as e:
            print(f"MySQL Error: {e}")
            flash(f'Database error: {str(e)}', 'error')
            return redirect(url_for('registration'))

        except Exception as e:
            print(f"General Error: {e}")
            flash(f'An error occurred: {str(e)}', 'error')
            return redirect(url_for('registration'))

@app.route('/admin-login')
def admin_login():
    return render_template('admin-login.html')

@app.route('/admin-auth', methods=['POST'])
def admin_auth():
    if request.method == 'POST':
        try:
            username = request.form['username']
            password = request.form['password']

            conn = get_db_connection()
            cur = conn.cursor(dictionary=True)

            # Check if admin exists
            cur.execute("SELECT * FROM Admin WHERE name = %s", [username])
            admin = cur.fetchone()
            cur.close()
            conn.close()

            if admin and admin['password'] == password:  # For now, using plain password comparison
                session['admin_logged_in'] = True
                flash('Login successful!', 'success')
                return redirect(url_for('admin_dashboard'))
            else:
                flash('Invalid username or password!', 'error')
                return redirect(url_for('admin_login'))

        except Exception as e:
            flash(f'An error occurred during login: {str(e)}', 'error')
            return redirect(url_for('admin_login'))

@app.route('/admin-dashboard')
def admin_dashboard():
    if not session.get('admin_logged_in'):
        flash('Please login first!', 'error')
        return redirect(url_for('admin_login'))

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM Students ORDER BY rollNumber")
        registrations = cur.fetchall()
        cur.close()
        conn.close()
        return render_template('admin-dashboard.html', registrations=registrations)

    except Exception as e:
        flash(f'Error fetching registrations: {str(e)}', 'error')
        return redirect(url_for('admin_login'))

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully!', 'success')
    return redirect(url_for('registration'))

@app.route('/export-excel')
def export_excel():
    if not session.get('admin_logged_in'):
        flash('Please login first!', 'error')
        return redirect(url_for('admin_login'))

    try:
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Registered Students"

        # Add headers
        headers = ['Name', 'Roll Number', 'Register Number', 'Email', 'Department', 'Registration Date', 'Last Updated']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Get data from database
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT name, rollNumber, registerNumber, email, department, created_at, updated_at FROM Students ORDER BY name")
        students = cur.fetchall()
        cur.close()
        conn.close()

        # Add data to worksheet
        for row, student in enumerate(students, 2):
            ws.cell(row=row, column=1, value=student['name'])
            ws.cell(row=row, column=2, value=student['rollNumber'])
            ws.cell(row=row, column=3, value=student['registerNumber'])
            ws.cell(row=row, column=4, value=student['email'])
            ws.cell(row=row, column=5, value=student['department'])
            ws.cell(row=row, column=6, value=student['created_at'].strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=7, value=student['updated_at'].strftime('%Y-%m-%d %H:%M:%S'))

        # Create a BytesIO object to store the Excel file
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate filename with current date
        current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'registered_students_{current_date}.xlsx'

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        flash(f'Error exporting data: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))

if __name__ == '__main__':
    # Test database connection on startup
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT 1")
        cur.fetchall()  # Fetch the result to clear it
        cur.close()
        conn.close()
        print("Database connection test successful on startup")
    except Exception as e:
        print(f"Database connection test failed on startup: {e}")

    app.run(host='0.0.0.0', debug=True, port=5000)